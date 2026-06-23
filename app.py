from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import time
import uuid
import urllib.request
import urllib.error
from cgi import FieldStorage
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("VOC_DATA_DIR", ROOT / "data"))
EXPORT_DIR = Path(os.environ.get("VOC_EXPORT_DIR", ROOT / "exports"))
STATIC_DIR = ROOT / "static"
PROJECTS_FILE = DATA_DIR / "projects.json"

DEEPSEEK_BASE_URL = os.environ.get("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
DEFAULT_FAST_MODEL = os.environ.get("DEEPSEEK_FAST_MODEL", "deepseek-v4-flash")
DEFAULT_ACCURATE_MODEL = os.environ.get("DEEPSEEK_ACCURATE_MODEL", "deepseek-v4-pro")
ACCESS_CODE = os.environ.get("VOC_ACCESS_CODE", "").strip()

USAGE_MARKS = ["产品表现", "用户/关系", "场景/目的", "使用工具/搭配对象", "购买行为/态度", "泛化评价"]
SENTIMENTS = ["P", "N", "M", "事实提及", "Drop"]

SKILL_RULE = """你是 VOC 语义打标专家。
任务是逐条完整阅读 Review 原文，提取最小语义标签。
禁止关键词匹配式打标。关键词只能提醒你检查语义，不能决定标签。
英文原文是主证据，中文翻译只能辅助。
最小语义标签是 Review 有效信息片段的精炼中文短句，不是完整翻译，也不是维度归纳。
不能加入原文没有的信息。不能过度泛化。必须保留关键对象、场景、工具、用户、产品表现和结果。
服务、物流、包装、二手到手状态等外部因素不要误标为产品表现；可进入 drop_records 或其他用途标记。
输出必须是严格 JSON，不要输出 Markdown。"""


def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    if not PROJECTS_FILE.exists():
        PROJECTS_FILE.write_text("[]", encoding="utf-8")


def read_json(path: Path, default):
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def write_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def load_projects():
    ensure_dirs()
    return read_json(PROJECTS_FILE, [])


def save_projects(projects):
    write_json(PROJECTS_FILE, projects)


def project_dir(project_id: str) -> Path:
    return DATA_DIR / project_id


def get_project(project_id: str):
    for p in load_projects():
        if p["id"] == project_id:
            return p
    return None


def update_project(project_id: str, patch: dict):
    projects = load_projects()
    for i, p in enumerate(projects):
        if p["id"] == project_id:
            projects[i] = {**p, **patch, "updated_at": now_iso()}
            save_projects(projects)
            return projects[i]
    return None


def delete_project(project_id: str):
    projects = load_projects()
    project = next((p for p in projects if p.get("id") == project_id), None)
    if not project:
        return None
    save_projects([p for p in projects if p.get("id") != project_id])
    src = project_dir(project_id)
    if src.exists():
        deleted_dir = DATA_DIR / "_deleted_projects"
        deleted_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        dst = deleted_dir / f"{stamp}-{project_id}"
        shutil.move(str(src), str(dst))
        write_json(dst / "deleted_project_meta.json", {"deleted_at": now_iso(), "project": project})
    return project


def body_json(handler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length) if length else b"{}"
    if not raw:
        return {}
    return json.loads(raw.decode("utf-8"))


def send_json(handler, data, status=200):
    payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(payload)))
    handler.end_headers()
    handler.wfile.write(payload)


def send_html(handler, html: str, status=200):
    payload = html.encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "text/html; charset=utf-8")
    handler.send_header("Content-Length", str(len(payload)))
    handler.end_headers()
    handler.wfile.write(payload)


def send_file(handler, path: Path, content_type: str):
    if not path.exists():
        send_json(handler, {"error": "file_not_found"}, 404)
        return
    data = path.read_bytes()
    handler.send_response(200)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Content-Length", str(len(data)))
    handler.send_header("Content-Disposition", f'attachment; filename="{path.name}"')
    handler.end_headers()
    handler.wfile.write(data)


def has_access(handler) -> bool:
    if not ACCESS_CODE:
        return True
    cookie = handler.headers.get("Cookie", "")
    return "voc_access=1" in cookie


def login_page():
    return """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>VOC 工具访问</title>
  <style>
    body{margin:0;min-height:100vh;display:grid;place-items:center;background:#f5f7fa;font-family:"Microsoft YaHei",Arial,sans-serif;color:#1f2933}
    form{width:min(420px,90vw);background:white;border:1px solid #d8e0e8;border-radius:8px;padding:24px;box-shadow:0 10px 30px rgba(31,78,121,.08)}
    h1{font-size:20px;margin:0 0 8px}p{color:#667085;font-size:13px;line-height:1.6;margin:0 0 18px}
    input,button{width:100%;box-sizing:border-box;border-radius:6px;font:inherit;padding:10px 12px}
    input{border:1px solid #cbd5e1;margin-bottom:12px}button{border:1px solid #1f4e79;background:#1f4e79;color:white;cursor:pointer}
    .error{color:#b42318;font-size:13px;min-height:20px;margin-top:10px}
  </style>
</head>
<body>
  <form id="loginForm">
    <h1>VOC 语义打标工具</h1>
    <p>请输入访问码。DeepSeek API Key 已由服务端配置，页面不会展示或保存密钥。</p>
    <input id="code" type="password" autocomplete="current-password" placeholder="访问码" autofocus />
    <button>进入工具</button>
    <div id="error" class="error"></div>
  </form>
  <script>
    document.getElementById('loginForm').addEventListener('submit', async (e) => {
      e.preventDefault();
      const res = await fetch('/api/access', {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body:JSON.stringify({code:document.getElementById('code').value})
      });
      if (res.ok) location.href = '/';
      else document.getElementById('error').textContent = '访问码不正确';
    });
  </script>
</body>
</html>"""


HEADER_SCAN_ROWS = 20
TEXT_SAMPLE_ROWS = 80

ID_HEADERS = ["Review序号", "序号", "review_id", "ReviewID", "review id", "id", "评论ID", "评价ID"]
TITLE_HEADERS = ["标题", "title", "review_title", "review title", "reviewtitle", "评论标题", "评价标题"]
BODY_HEADERS = [
    "正文",
    "review原文",
    "review 原文",
    "review",
    "review body",
    "review text",
    "review_body",
    "review_text",
    "reviewbody",
    "reviewtext",
    "body",
    "content",
    "reviewcontent",
    "评论内容",
    "评论正文",
    "评论原文",
    "评价内容",
    "评价正文",
    "英文原文",
    "英文review",
    "review英文原文",
    "客户评论",
]
CN_HEADERS = ["Review中文翻译", "review-中文翻译", "review 中文翻译", "中文翻译", "translation", "review translation"]
STAR_HEADERS = ["星级", "rating", "ratings", "stars", "star rating", "score", "评分"]
ASIN_HEADERS = ["ASIN", "asin"]
MODEL_HEADERS = ["型号", "变体", "variant", "variation", "model", "产品型号", "variant name", "sku"]
LINK_HEADERS = ["Review链接", "链接", "reviewlink", "review link", "review_url", "review url", "url", "permalink"]


def normalize_header(value) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_\-:/\\|,.;()（）【】\[\]{}<>]+", "", text)


def find_header(headers, candidates):
    normalized = [normalize_header(h) for h in headers]
    for cand in candidates:
        key = normalize_header(cand)
        if key in normalized:
            return normalized.index(key)
    return None


def parse_reviews_from_xlsx(data: bytes):
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    best_reviews = []
    best_info = {"file_type": "xlsx", "sheets": [ws.title for ws in wb.worksheets]}
    best_score = -1
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        reviews, info, score = parse_table_rows(rows, {"file_type": "xlsx", "sheet": ws.title})
        if score > best_score or (score == best_score and len(reviews) > len(best_reviews)):
            best_reviews = reviews
            best_info = info
            best_score = score
    return best_reviews, best_info


def parse_reviews_from_csv(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    reviews, info, _score = parse_table_rows(rows, {"file_type": "csv"})
    return reviews, info


def cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx] or "").strip()


def parse_table_rows(rows, base_info):
    if not rows:
        return [], {**base_info, "message": "文件里没有可读取的行"}, 0

    best_reviews = []
    best_info = {**base_info, "message": "没有识别到 Review 正文列"}
    best_score = -1
    limit = min(len(rows), HEADER_SCAN_ROWS)
    for header_idx in range(limit):
        headers = [str(x or "").strip() for x in rows[header_idx]]
        if not any(headers):
            continue
        data_rows = rows[header_idx + 1 :]
        explicit_score = header_match_score(headers)
        inferred_body_idx = None
        if find_header(headers, BODY_HEADERS) is None and find_header(headers, TITLE_HEADERS) is None:
            inferred_body_idx = infer_body_column(headers, data_rows)
        reviews = rows_to_reviews(headers, data_rows, inferred_body_idx)
        if not reviews:
            continue
        score = explicit_score * 1000 + min(len(reviews), 999)
        if inferred_body_idx is not None:
            score += 100
        if score > best_score or (score == best_score and len(reviews) > len(best_reviews)):
            best_reviews = reviews
            best_info = {
                **base_info,
                "header_row": header_idx + 1,
                "headers_preview": [h for h in headers if h][:20],
                "detected_columns": detected_columns(headers, data_rows, inferred_body_idx),
                "body_column_was_inferred": inferred_body_idx is not None,
            }
            best_score = score
    return best_reviews, best_info, max(best_score, 0)


def header_match_score(headers):
    score = 0
    if find_header(headers, BODY_HEADERS) is not None:
        score += 10
    if find_header(headers, TITLE_HEADERS) is not None:
        score += 4
    if find_header(headers, ID_HEADERS) is not None:
        score += 2
    if find_header(headers, STAR_HEADERS) is not None:
        score += 1
    if find_header(headers, MODEL_HEADERS) is not None:
        score += 1
    return score


def infer_body_column(headers, data_rows):
    candidates = []
    max_cols = max([len(headers)] + [len(r) for r in data_rows[:TEXT_SAMPLE_ROWS]] or [0])
    for idx in range(max_cols):
        header = normalize_header(headers[idx] if idx < len(headers) else "")
        if header in {normalize_header(x) for x in ID_HEADERS + STAR_HEADERS + ASIN_HEADERS + MODEL_HEADERS + LINK_HEADERS}:
            continue
        values = [cell(row, idx) for row in data_rows[:TEXT_SAMPLE_ROWS] if cell(row, idx)]
        if len(values) < 2:
            continue
        numeric_like = sum(1 for value in values if re.fullmatch(r"[\d.]+", value))
        url_like = sum(1 for value in values if value.startswith(("http://", "https://")))
        if numeric_like >= len(values) * 0.7 or url_like >= len(values) * 0.5:
            continue
        long_values = [value for value in values if len(value) >= 20]
        language_values = [value for value in values if re.search(r"[A-Za-z]{3,}|[\u4e00-\u9fff]{2,}", value)]
        if len(long_values) < 2 or len(language_values) < 2:
            continue
        avg_len = sum(len(value) for value in values) / len(values)
        max_len = max(len(value) for value in values)
        score = len(long_values) * 10 + avg_len + max_len / 5
        candidates.append((score, idx))
    candidates.sort(reverse=True)
    if not candidates:
        return None
    if len(candidates) == 1 or candidates[0][0] >= candidates[1][0] * 1.2:
        return candidates[0][1]
    return None


def review_column_indexes(headers, data_rows, inferred_body_idx=None):
    body_idx = find_header(headers, BODY_HEADERS)
    return {
        "review_id": find_header(headers, ID_HEADERS),
        "title": find_header(headers, TITLE_HEADERS),
        "body": body_idx if body_idx is not None else inferred_body_idx,
        "translation_zh": find_header(headers, CN_HEADERS),
        "star": find_header(headers, STAR_HEADERS),
        "asin": find_header(headers, ASIN_HEADERS),
        "model": find_header(headers, MODEL_HEADERS),
        "link": find_header(headers, LINK_HEADERS),
    }


def detected_columns(headers, data_rows, inferred_body_idx=None):
    indexes = review_column_indexes(headers, data_rows, inferred_body_idx)
    result = {}
    for key, idx in indexes.items():
        result[key] = headers[idx] if idx is not None and idx < len(headers) else ""
    return result


def rows_to_reviews(headers, data_rows, inferred_body_idx=None):
    indexes = review_column_indexes(headers, data_rows, inferred_body_idx)
    id_idx = indexes["review_id"]
    title_idx = indexes["title"]
    body_idx = indexes["body"]
    cn_idx = indexes["translation_zh"]
    star_idx = indexes["star"]
    asin_idx = indexes["asin"]
    model_idx = indexes["model"]
    link_idx = indexes["link"]

    reviews = []
    for n, row in enumerate(data_rows, start=1):
        title = cell(row, title_idx)
        body = cell(row, body_idx)
        if title and body:
            original = f"[{title}]\n{body}"
        else:
            original = body or title
        if not original.strip():
            continue
        reviews.append(
            {
                "review_id": cell(row, id_idx) or f"R{n:04d}",
                "seq": n,
                "asin": cell(row, asin_idx),
                "model": cell(row, model_idx),
                "star": cell(row, star_idx),
                "review_original": original,
                "review_translation_zh": cell(row, cn_idx),
                "review_link": cell(row, link_idx),
            }
        )
    return reviews


def estimate_tokens(text: str) -> int:
    words = len(re.findall(r"[A-Za-z0-9]+", text or ""))
    return max(int(words * 1.3), int(len(text or "") / 4), 1)


def project_stats(project_id: str):
    reviews = read_json(project_dir(project_id) / "reviews.json", [])
    batches = read_json(project_dir(project_id) / "batches.json", [])
    atomic = read_json(project_dir(project_id) / "atomic_results.json", [])
    final_labels = read_json(project_dir(project_id) / "final_labels.json", [])
    locked = read_locked_rules(project_id)
    analysis = read_json(project_dir(project_id) / "analysis_summary.json", {})
    review_tokens = sum(estimate_tokens(r.get("review_original", "")) for r in reviews)
    atomic_need_review = sum(1 for x in atomic if x.get("need_review"))
    final_need_review = sum(1 for x in final_labels if x.get("need_review"))
    return {
        "reviews": len(reviews),
        "batches": len(batches),
        "atomic_records": sum(len(x.get("atomic_tags", [])) for x in atomic),
        "drop_records": sum(len(x.get("drop_records", [])) for x in atomic),
        "need_review": atomic_need_review + final_need_review,
        "atomic_need_review": atomic_need_review,
        "final_need_review": final_need_review,
        "final_labeled": len(final_labels),
        "locked_decision_dimensions": len(locked.get("decision_dimensions", []) or []),
        "locked_context_fields": len(locked.get("context_fields", []) or []),
        "analysis_ready": bool(analysis.get("generated_at")),
        "analysis_records": len(analysis.get("decision_subdimensions", []) or []) + len(analysis.get("context_subdimensions", []) or []),
        "approx_review_tokens": review_tokens,
    }


def archive_and_remove(project_id: str, filenames: list[str]) -> None:
    base = project_dir(project_id)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    archive_dir = base / "archive" / stamp
    moved = False
    for name in filenames:
        path = base / name
        if not path.exists():
            continue
        archive_dir.mkdir(parents=True, exist_ok=True)
        path.replace(archive_dir / name)
        moved = True
    if moved:
        write_json(archive_dir / "archive_note.json", {"archived_at": now_iso(), "reason": "new upstream data invalidated downstream outputs"})


def reset_after_reviews_upload(project_id: str) -> None:
    archive_and_remove(
        project_id,
        [
            "reviews.json",
            "batches.json",
            "atomic_results.json",
            "dimension_candidates.json",
            "dimension_model.json",
            "locked_dimensions.json",
            "final_labels.json",
            "analysis_summary.json",
        ],
    )


def reset_after_atomic_change(project_id: str) -> None:
    archive_and_remove(project_id, ["dimension_model.json", "locked_dimensions.json", "final_labels.json", "analysis_summary.json"])


def reset_after_dimension_change(project_id: str) -> None:
    archive_and_remove(project_id, ["final_labels.json", "analysis_summary.json"])


def is_locked_rules(data: dict) -> bool:
    return bool(data.get("locked_at") and ((data.get("decision_dimensions") or []) or (data.get("context_fields") or [])))


def read_locked_rules(project_id: str) -> dict:
    data = read_json(project_dir(project_id) / "locked_dimensions.json", {})
    return data if is_locked_rules(data) else {}


def make_batches(project_id: str, batch_size: int):
    reviews = read_json(project_dir(project_id) / "reviews.json", [])
    batches = []
    for i in range(0, len(reviews), batch_size):
        chunk = reviews[i : i + batch_size]
        batches.append(
            {
                "id": f"B{len(batches)+1:03d}",
                "start_seq": chunk[0]["seq"],
                "end_seq": chunk[-1]["seq"],
                "review_count": len(chunk),
                "status": "pending",
                "model": "",
                "started_at": "",
                "finished_at": "",
                "error": "",
            }
        )
    write_json(project_dir(project_id) / "batches.json", batches)
    return batches


def batch_reviews(project_id: str, batch_id: str):
    reviews = read_json(project_dir(project_id) / "reviews.json", [])
    batches = read_json(project_dir(project_id) / "batches.json", [])
    batch = next((b for b in batches if b["id"] == batch_id), None)
    if not batch:
        return None, []
    selected = [r for r in reviews if batch["start_seq"] <= r["seq"] <= batch["end_seq"]]
    return batch, selected


def set_batch_status(project_id: str, batch_id: str, **patch):
    path = project_dir(project_id) / "batches.json"
    batches = read_json(path, [])
    for b in batches:
        if b["id"] == batch_id:
            b.update(patch)
    write_json(path, batches)


def extract_json_from_text(text: str):
    raw = (text or "").strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?", "", raw, flags=re.I).strip()
        raw = re.sub(r"```$", "", raw).strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw, re.S)
        if match:
            return json.loads(match.group(0))
        raise


def deepseek_chat(api_key: str, model: str, messages: list[dict], max_tokens=12000, temperature=0.1):
    payload = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
        "response_format": {"type": "json_object"},
    }
    req = urllib.request.Request(
        f"{DEEPSEEK_BASE_URL.rstrip('/')}/chat/completions",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=240) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"DeepSeek HTTP {e.code}: {detail[:1000]}")
    except urllib.error.URLError as e:
        raise RuntimeError(f"DeepSeek connection error: {e}")

    content = data["choices"][0]["message"]["content"]
    usage = data.get("usage", {})
    return extract_json_from_text(content), usage


def atomic_prompt(project: dict, reviews: list[dict]) -> list[dict]:
    project_block = {
        "product_name": project.get("name", ""),
        "product_category": project.get("category", ""),
        "analysis_goal": "先发现式拆解最小语义标签，不生成最终维度。",
        "usage_marks_allowed": USAGE_MARKS,
        "sentiments_allowed": SENTIMENTS,
    }
    task = {
        "task": "extract_atomic_tags",
        "requirements": [
            "逐条完整阅读 review_original。",
            "一条 review 可以拆多个 atomic_tags。",
            "atomic_tag_zh 必须是中文短句；品牌、型号、缩写、专有名词可保留英文。",
            "evidence_original 必须是原文短证据，不要整段复制。",
            "usage_marks 可以多选，只能来自枚举。",
            "sentiment 只能是 P、N、M、事实提及。Drop 内容放 drop_records。",
            "不要生成购买决策维度；不要做聚类。",
            "不确定时 need_review=true，并写 review_flags。",
        ],
        "output_schema": {
            "reviews": [
                {
                    "review_id": "string",
                    "atomic_tags": [
                        {
                            "atomic_tag_zh": "string",
                            "evidence_original": "string",
                            "sentiment": "P|N|M|事实提及",
                            "usage_marks": ["产品表现"],
                            "confidence": 0.0,
                        }
                    ],
                    "drop_records": [{"text": "string", "drop_reason": "string"}],
                    "review_flags": ["string"],
                    "need_review": False,
                }
            ]
        },
        "reviews": [
            {
                "review_id": r["review_id"],
                "star": r.get("star", ""),
                "model": r.get("model", ""),
                "review_original": r.get("review_original", ""),
                "review_translation_zh": r.get("review_translation_zh", ""),
            }
            for r in reviews
        ],
    }
    return [
        {"role": "system", "content": SKILL_RULE + "\n必须输出 json object。"},
        {"role": "user", "content": "项目配置 json：\n" + json.dumps(project_block, ensure_ascii=False)},
        {"role": "user", "content": "任务 json：\n" + json.dumps(task, ensure_ascii=False)},
    ]


def validate_atomic_result(result: dict, expected_ids: set[str]):
    errors = []
    rows = result.get("reviews")
    if not isinstance(rows, list):
        return ["missing reviews array"]
    seen = set()
    for item in rows:
        rid = str(item.get("review_id", "")).strip()
        if not rid:
            errors.append("missing review_id")
            continue
        seen.add(rid)
        if rid not in expected_ids:
            errors.append(f"unexpected review_id {rid}")
        for tag in item.get("atomic_tags", []) or []:
            if not tag.get("atomic_tag_zh"):
                errors.append(f"{rid}: empty atomic_tag_zh")
            if not tag.get("evidence_original"):
                errors.append(f"{rid}: empty evidence_original")
            if tag.get("sentiment") not in ["P", "N", "M", "事实提及"]:
                errors.append(f"{rid}: invalid sentiment {tag.get('sentiment')}")
            marks = tag.get("usage_marks") or []
            if not isinstance(marks, list) or any(m not in USAGE_MARKS for m in marks):
                errors.append(f"{rid}: invalid usage_marks {marks}")
    missing = expected_ids - seen
    if missing:
        errors.append(f"missing review ids: {', '.join(sorted(missing)[:10])}")
    return errors


def mock_atomic_extract(reviews: list[dict]):
    rows = []
    for r in reviews:
        text = r.get("review_original", "")
        lowered = text.lower()
        tags = []
        drops = []
        if any(k in lowered for k in ["great", "good", "love", "perfect", "excellent"]):
            tags.append(
                {
                    "atomic_tag_zh": "整体评价正向但需人工确认具体原因",
                    "evidence_original": text[:120],
                    "sentiment": "P",
                    "usage_marks": ["泛化评价"],
                    "confidence": 0.35,
                }
            )
        if any(k in lowered for k in ["broken", "doesn't work", "not work", "failed", "bad"]):
            tags.append(
                {
                    "atomic_tag_zh": "出现无法正常使用或故障反馈",
                    "evidence_original": text[:120],
                    "sentiment": "N",
                    "usage_marks": ["产品表现"],
                    "confidence": 0.35,
                }
            )
        if any(k in lowered for k in ["package", "shipping", "delivered", "seller", "customer service", "warranty"]):
            drops.append({"text": text[:120], "drop_reason": "疑似售后、履约或包装信息，需人工确认是否进入产品表现"})
        rows.append(
            {
                "review_id": r["review_id"],
                "atomic_tags": tags,
                "drop_records": drops,
                "review_flags": ["模拟运行结果，不作为正式打标"] if tags or drops else ["未识别出高置信最小语义标签"],
                "need_review": True,
            }
        )
    return {"reviews": rows}, {"mock": True}


def merge_atomic_results(project_id: str, batch_id: str, result: dict, usage: dict):
    path = project_dir(project_id) / "atomic_results.json"
    existing = read_json(path, [])
    batch_ids = {r.get("batch_id") for r in existing}
    if batch_id in batch_ids:
        existing = [r for r in existing if r.get("batch_id") != batch_id]
    for item in result.get("reviews", []):
        item["batch_id"] = batch_id
        item["usage"] = usage
        existing.append(item)
    write_json(path, existing)
    reset_after_atomic_change(project_id)


def propose_dimensions(project_id: str):
    atomic = read_json(project_dir(project_id) / "atomic_results.json", [])
    product_counter = {}
    context_counter = {}
    for row in atomic:
        for tag in row.get("atomic_tags", []) or []:
            text = tag.get("atomic_tag_zh", "")
            marks = tag.get("usage_marks", []) or []
            if "产品表现" in marks:
                product_counter[text] = product_counter.get(text, 0) + 1
            if any(m in marks for m in ["用户/关系", "场景/目的", "使用工具/搭配对象", "购买行为/态度"]):
                context_counter[text] = context_counter.get(text, 0) + 1
    product = sorted(product_counter.items(), key=lambda x: (-x[1], x[0]))[:80]
    context = sorted(context_counter.items(), key=lambda x: (-x[1], x[0]))[:80]
    candidates = {
        "note": "基于最小语义标签频次生成的候选池。维度定义仍需要 DeepSeek 语义归类和人工锁定。",
        "product_atomic_pool": [{"atomic_tag": k, "count": v} for k, v in product],
        "context_atomic_pool": [{"atomic_tag": k, "count": v} for k, v in context],
    }
    write_json(project_dir(project_id) / "dimension_candidates.json", candidates)
    return candidates


def dimension_model_prompt(project: dict, candidates: dict, atomic: list[dict]) -> list[dict]:
    compact_atomic = []
    for row in atomic[:350]:
        for tag in (row.get("atomic_tags", []) or [])[:8]:
            compact_atomic.append(
                {
                    "review_id": row.get("review_id", ""),
                    "atomic_tag_zh": tag.get("atomic_tag_zh", ""),
                    "sentiment": tag.get("sentiment", ""),
                    "usage_marks": tag.get("usage_marks", []),
                    "evidence_original": tag.get("evidence_original", ""),
                }
            )
    task = {
        "task": "build_dimension_draft_from_atomic_tags",
        "product": {
            "name": project.get("name", ""),
            "category": project.get("category", ""),
            "description": project.get("description", ""),
        },
        "candidates": candidates,
        "atomic_evidence_sample": compact_atomic,
        "requirements": [
            "不要按关键词机械聚类，要理解最小语义标签背后的购买决策问题。",
            "产品购买决策维度必须是买家会用来判断是否购买/留下/退货的一级问题。",
            "Context 字段必须是人群、场景、用途、购买路径、使用阻碍等背景信息，不要混入产品性能。",
            "维度数量不要过多。产品购买决策维度建议 6-10 个，Context 字段建议 8-14 个。",
            "每个维度都要写清楚 P/N/M/0 的边界，方便后续逐条打标。",
            "输出中文。不要写 Listing 文案，只写分析维度定义。",
        ],
        "output_schema": {
            "decision_dimensions": [
                {
                    "name_zh": "string",
                    "definition_zh": "string",
                    "decision_question_zh": "string",
                    "p_rule_zh": "string",
                    "n_rule_zh": "string",
                    "m_rule_zh": "string",
                    "zero_rule_zh": "string",
                    "boundary_zh": "string",
                    "source_atomic_tags": ["string"],
                    "listing_use_zh": "string",
                }
            ],
            "context_fields": [
                {
                    "name_zh": "string",
                    "definition_zh": "string",
                    "evidence_required_zh": "string",
                    "boundary_zh": "string",
                    "source_atomic_tags": ["string"],
                    "analysis_use_zh": "string",
                }
            ],
            "overflow_or_other": [{"theme_zh": "string", "reason_zh": "string"}],
            "need_human_decisions": ["string"],
        },
    }
    return [
        {"role": "system", "content": SKILL_RULE + "\n你现在只生成可讨论的维度草案，必须输出 json object。"},
        {"role": "user", "content": json.dumps(task, ensure_ascii=False)},
    ]


def mock_dimension_model(candidates: dict):
    product = candidates.get("product_atomic_pool", [])[:8]
    context = candidates.get("context_atomic_pool", [])[:10]
    return {
        "decision_dimensions": [
            {
                "name_zh": f"待确认产品维度{i+1}",
                "definition_zh": f"由最小语义标签“{item.get('atomic_tag', '')}”触发的候选维度，需要人工改名和补边界。",
                "decision_question_zh": "这个产品表现是否会影响买家购买或留用判断？",
                "p_rule_zh": "原文明确表达该表现满足或优于预期。",
                "n_rule_zh": "原文明确表达该表现失败、限制使用或低于预期。",
                "m_rule_zh": "只客观提到该表现，没有明确好坏。",
                "zero_rule_zh": "未提及或不能语义支持该维度。",
                "boundary_zh": "模拟草案，仅用于检查流程，不可直接作为最终规则。",
                "source_atomic_tags": [item.get("atomic_tag", "")],
                "listing_use_zh": "确认后可用于卖点排序、风险说明或图片信息层级。",
            }
            for i, item in enumerate(product)
        ],
        "context_fields": [
            {
                "name_zh": f"待确认Context{i+1}",
                "definition_zh": f"由最小语义标签“{item.get('atomic_tag', '')}”触发的候选 Context。",
                "evidence_required_zh": "必须有原文明确背景证据。",
                "boundary_zh": "模拟草案，仅用于检查流程，不可直接作为最终规则。",
                "source_atomic_tags": [item.get("atomic_tag", "")],
                "analysis_use_zh": "确认后可用于理解用户画像、使用场景或购买路径。",
            }
            for i, item in enumerate(context)
        ],
        "overflow_or_other": [],
        "need_human_decisions": ["模拟运行结果，不作为正式维度。"],
    }


def generate_dimension_model(project_id: str, api_key: str, model: str, mock: bool):
    project = get_project(project_id)
    atomic = read_json(project_dir(project_id) / "atomic_results.json", [])
    candidates = read_json(project_dir(project_id) / "dimension_candidates.json", {}) or propose_dimensions(project_id)
    if mock:
        draft, usage = mock_dimension_model(candidates), {"mock": True}
    else:
        draft, usage = deepseek_chat(api_key, model, dimension_model_prompt(project, candidates, atomic), max_tokens=16000, temperature=0.15)
    draft["model"] = model
    draft["usage"] = usage
    draft["generated_at"] = now_iso()
    archive_and_remove(project_id, ["locked_dimensions.json", "final_labels.json", "analysis_summary.json"])
    write_json(project_dir(project_id) / "dimension_model.json", draft)
    update_project(project_id, {"stage": "dimension_draft_ready"})
    return draft


def rules_for_project(project_id: str):
    locked = read_locked_rules(project_id)
    if locked:
        return locked
    return read_json(project_dir(project_id) / "dimension_model.json", {})


def atomic_for_review(project_id: str, review_ids: set[str]):
    atomic = read_json(project_dir(project_id) / "atomic_results.json", [])
    return [row for row in atomic if row.get("review_id") in review_ids]


def atomic_tag_review_index(atomic: list[dict]):
    index = {}
    for row in atomic:
        rid = row.get("review_id", "")
        for tag in row.get("atomic_tags", []) or []:
            text = tag.get("atomic_tag_zh", "")
            if not text:
                continue
            index.setdefault(text, set()).add(rid)
    return index


def rule_source_stats(project_id: str):
    reviews = read_json(project_dir(project_id) / "reviews.json", [])
    atomic = read_json(project_dir(project_id) / "atomic_results.json", [])
    rules = rules_for_project(project_id)
    tag_index = atomic_tag_review_index(atomic)
    total = max(len(reviews), 1)

    def stats_for(items):
        rows = []
        for item in items:
            source_tags = item.get("source_atomic_tags", []) or []
            review_ids = set()
            for tag in source_tags:
                review_ids.update(tag_index.get(tag, set()))
            rows.append(
                {
                    "name_zh": item.get("name_zh", ""),
                    "source_tag_count": len(source_tags),
                    "mention_count": len(review_ids),
                    "mention_rate": round(len(review_ids) / total, 4),
                    "sort_basis": "按来源最小语义标签覆盖的 Review 数降序",
                }
            )
        return sorted(rows, key=lambda x: (-x["mention_count"], x["name_zh"]))

    return {
        "decision": stats_for(rules.get("decision_dimensions", []) or []),
        "context": stats_for(rules.get("context_fields", []) or []),
    }


def final_label_stats(project_id: str):
    reviews = read_json(project_dir(project_id) / "reviews.json", [])
    final_labels = read_json(project_dir(project_id) / "final_labels.json", [])
    rules = rules_for_project(project_id)
    total = max(len(reviews), 1)
    decision = []
    for item in rules.get("decision_dimensions", []) or []:
        name = item.get("name_zh", "")
        counts = {"P": 0, "N": 0, "M": 0}
        review_ids = []
        for row in final_labels:
            val = (row.get("dimensions", {}) or {}).get(name, {})
            t = val.get("T", "0")
            if t in counts:
                counts[t] += 1
                review_ids.append(row.get("review_id", ""))
        mention_count = sum(counts.values())
        decision.append(
            {
                "name_zh": name,
                "mention_count": mention_count,
                "mention_rate": round(mention_count / total, 4),
                "p_count": counts["P"],
                "n_count": counts["N"],
                "m_count": counts["M"],
                "sort_basis": "按最终打标 T 不为 0 的 Review 数降序",
            }
        )
    context = []
    for item in rules.get("context_fields", []) or []:
        name = item.get("name_zh", "")
        review_ids = []
        for row in final_labels:
            val = (row.get("context", {}) or {}).get(name, {})
            value = str(val.get("value_zh", "")).strip()
            if value and value not in ["未提及", "无", "0"]:
                review_ids.append(row.get("review_id", ""))
        context.append(
            {
                "name_zh": name,
                "mention_count": len(review_ids),
                "mention_rate": round(len(review_ids) / total, 4),
                "sort_basis": "按最终打标有明确 Context 值的 Review 数降序",
            }
        )
    return {
        "decision": sorted(decision, key=lambda x: (-x["mention_count"], x["name_zh"])),
        "context": sorted(context, key=lambda x: (-x["mention_count"], x["name_zh"])),
    }


def project_dimension_stats(project_id: str):
    return {
        "source": rule_source_stats(project_id),
        "final": final_label_stats(project_id),
    }


def need_review_items(project_id: str, limit: int = 80):
    reviews = read_json(project_dir(project_id) / "reviews.json", [])
    review_map = {r.get("review_id"): r for r in reviews}
    items = []
    for row in read_json(project_dir(project_id) / "atomic_results.json", []):
        if not row.get("need_review"):
            continue
        evidence = ""
        tags = row.get("atomic_tags", []) or []
        drops = row.get("drop_records", []) or []
        if tags:
            evidence = tags[0].get("evidence_original", "")
        elif drops:
            evidence = drops[0].get("text", "")
        rid = row.get("review_id", "")
        items.append(
            {
                "stage": "最小语义标签",
                "review_id": rid,
                "reason": "；".join(row.get("review_flags", []) or []) or "低置信或边界不清",
                "evidence": evidence,
                "review_original": (review_map.get(rid, {}) or {}).get("review_original", ""),
            }
        )
    for row in read_json(project_dir(project_id) / "final_labels.json", []):
        if not row.get("need_review"):
            continue
        rid = row.get("review_id", "")
        evidence = ""
        for val in (row.get("dimensions", {}) or {}).values():
            if val.get("R") and val.get("R") != "无":
                evidence = val.get("R")
                break
        if not evidence:
            for val in (row.get("context", {}) or {}).values():
                if val.get("evidence_zh") and val.get("evidence_zh") != "无":
                    evidence = val.get("evidence_zh")
                    break
        items.append(
            {
                "stage": "最终打标",
                "review_id": rid,
                "reason": "；".join(row.get("review_flags", []) or []) or "低置信、边界冲突或可能幻觉",
                "evidence": evidence,
                "review_original": (review_map.get(rid, {}) or {}).get("review_original", ""),
            }
        )
    return items[:limit]


def clean_value(value: str) -> str:
    text = str(value or "").strip()
    return text if text and text not in ["无", "未提及", "0"] else ""


def entry_point_for_decision(name: str, p_count: int, n_count: int, m_count: int) -> str:
    if n_count > p_count and n_count > 0:
        return f"优先检查“{name}”相关差评：Listing 需提前说明限制，产品侧需判断是否为高频缺陷。"
    if p_count > 0 and p_count >= n_count:
        return f"可把“{name}”作为卖点候选，用图片/要点展示真实使用收益，同时保留适用边界。"
    if m_count > 0:
        return f"用户有事实提及但倾向不强，适合补充规格说明或 FAQ，暂不作为核心卖点。"
    return "样本较少，先保留观察，不单独形成运营动作。"


def build_analysis_summary(project_id: str):
    reviews = read_json(project_dir(project_id) / "reviews.json", [])
    final_labels = read_json(project_dir(project_id) / "final_labels.json", [])
    rules = rules_for_project(project_id)
    total = max(len(reviews), 1)
    decision_rows = []
    context_rows = []

    for dim in rules.get("decision_dimensions", []) or []:
        name = dim.get("name_zh", "")
        buckets = {}
        for row in final_labels:
            item = (row.get("dimensions", {}) or {}).get(name, {})
            t = item.get("T", "0")
            if t == "0":
                continue
            value = clean_value(item.get("atomic_value_zh")) or clean_value(item.get("R")) or "未命名细分"
            bucket = buckets.setdefault(value, {"P": 0, "N": 0, "M": 0, "review_ids": [], "evidence": []})
            if t in ["P", "N", "M"]:
                bucket[t] += 1
            bucket["review_ids"].append(row.get("review_id", ""))
            if item.get("R") and item.get("R") != "无" and len(bucket["evidence"]) < 3:
                bucket["evidence"].append(item.get("R"))
        for value, bucket in buckets.items():
            count = len(set(bucket["review_ids"]))
            decision_rows.append(
                {
                    "parent_dimension": name,
                    "subdimension_zh": value,
                    "mention_count": count,
                    "mention_rate": round(count / total, 4),
                    "p_count": bucket["P"],
                    "n_count": bucket["N"],
                    "m_count": bucket["M"],
                    "optimization_entry_zh": entry_point_for_decision(name, bucket["P"], bucket["N"], bucket["M"]),
                    "evidence_examples": bucket["evidence"],
                    "review_ids": sorted(set(bucket["review_ids"]))[:20],
                }
            )

    for field in rules.get("context_fields", []) or []:
        name = field.get("name_zh", "")
        buckets = {}
        for row in final_labels:
            item = (row.get("context", {}) or {}).get(name, {})
            value = clean_value(item.get("value_zh"))
            if not value:
                continue
            bucket = buckets.setdefault(value, {"review_ids": [], "evidence": []})
            bucket["review_ids"].append(row.get("review_id", ""))
            if item.get("evidence_zh") and item.get("evidence_zh") != "无" and len(bucket["evidence"]) < 3:
                bucket["evidence"].append(item.get("evidence_zh"))
        for value, bucket in buckets.items():
            count = len(set(bucket["review_ids"]))
            context_rows.append(
                {
                    "context_field": name,
                    "subdimension_zh": value,
                    "mention_count": count,
                    "mention_rate": round(count / total, 4),
                    "analysis_use_zh": f"用于判断“{name}”下的核心用户/场景/路径，辅助 Listing 信息层级和运营人群判断。",
                    "evidence_examples": bucket["evidence"],
                    "review_ids": sorted(set(bucket["review_ids"]))[:20],
                }
            )

    decision_rows.sort(key=lambda x: (-x["mention_count"], x["parent_dimension"], x["subdimension_zh"]))
    context_rows.sort(key=lambda x: (-x["mention_count"], x["context_field"], x["subdimension_zh"]))
    summary = {
        "generated_at": now_iso(),
        "basis": "基于最终打标宽表中的原子属性/Context 值做可追溯统计；未做不可解释的关键词匹配。",
        "decision_subdimensions": decision_rows,
        "context_subdimensions": context_rows,
    }
    write_json(project_dir(project_id) / "analysis_summary.json", summary)
    update_project(project_id, {"stage": "analysis_ready"})
    return summary


def final_label_prompt(project: dict, rules: dict, reviews: list[dict], atomic_rows: list[dict]) -> list[dict]:
    task = {
        "task": "label_reviews_with_locked_dimensions",
        "product": {
            "name": project.get("name", ""),
            "category": project.get("category", ""),
            "description": project.get("description", ""),
        },
        "decision_dimensions": rules.get("decision_dimensions", []),
        "context_fields": rules.get("context_fields", []),
        "atomic_tags_reference": atomic_rows,
        "reviews": [
            {
                "review_id": r["review_id"],
                "star": r.get("star", ""),
                "model": r.get("model", ""),
                "review_original": r.get("review_original", ""),
                "review_translation_zh": r.get("review_translation_zh", ""),
            }
            for r in reviews
        ],
        "requirements": [
            "必须逐条完整阅读 review_original，最小语义标签只是辅助线索，不能替代原文判断。",
            "英文原文是主证据；已有中文翻译不可靠时必须忽略。",
            "每个购买决策维度都要输出 T=P/N/M/0；T=0 时 R 和 atomic_value_zh 都写“无”。",
            "R 必须是中文证据摘要，不要整段英文；atomic_value_zh 是比 R 更短的原子属性。",
            "Context 字段只输出原文明确支持的背景值；未提及写“未提及”。",
            "不要把售后、物流、包装、二手到手状态误标为产品性能或购买 Context。",
            "低置信、边界冲突、可能幻觉时 need_review=true 并写 flags。",
        ],
        "output_schema": {
            "reviews": [
                {
                    "review_id": "string",
                    "dimensions": {
                        "维度名": {
                            "T": "P|N|M|0",
                            "R": "中文证据摘要或无",
                            "atomic_value_zh": "中文原子属性或无",
                            "confidence": "high|medium|low",
                            "flags": ["string"],
                        }
                    },
                    "context": {
                        "Context名": {
                            "value_zh": "中文原子属性或未提及",
                            "evidence_zh": "中文证据摘要或无",
                            "confidence": "high|medium|low",
                            "flags": ["string"],
                        }
                    },
                    "other": {"T": "P|N|M|0", "R": "中文证据摘要或无"},
                    "need_review": False,
                    "review_flags": ["string"],
                }
            ]
        },
    }
    return [
        {"role": "system", "content": SKILL_RULE + "\n你现在做最终 review 宽表打标，必须输出 json object。"},
        {"role": "user", "content": json.dumps(task, ensure_ascii=False)},
    ]


def normalize_final_result(result: dict, expected_ids: set[str], rules: dict):
    errors = []
    rows = result.get("reviews")
    if not isinstance(rows, list):
        return {"reviews": []}, ["missing reviews array"]
    dim_names = [d.get("name_zh", "") for d in rules.get("decision_dimensions", []) if d.get("name_zh")]
    ctx_names = [c.get("name_zh", "") for c in rules.get("context_fields", []) if c.get("name_zh")]
    normalized = []
    seen = set()
    for item in rows:
        rid = str(item.get("review_id", "")).strip()
        if not rid:
            errors.append("missing review_id")
            continue
        seen.add(rid)
        if rid not in expected_ids:
            errors.append(f"unexpected review_id {rid}")
        dims = item.get("dimensions") if isinstance(item.get("dimensions"), dict) else {}
        for name in dim_names:
            val = dims.get(name) if isinstance(dims.get(name), dict) else {}
            t = val.get("T", "0")
            if t not in ["P", "N", "M", "0"]:
                errors.append(f"{rid}: invalid T {name}={t}")
                t = "0"
            dims[name] = {
                "T": t,
                "R": val.get("R") or ("无" if t == "0" else ""),
                "atomic_value_zh": val.get("atomic_value_zh") or ("无" if t == "0" else ""),
                "confidence": val.get("confidence", "medium"),
                "flags": val.get("flags", []) or [],
            }
        ctx = item.get("context") if isinstance(item.get("context"), dict) else {}
        for name in ctx_names:
            val = ctx.get(name) if isinstance(ctx.get(name), dict) else {}
            value = val.get("value_zh") or "未提及"
            ctx[name] = {
                "value_zh": value,
                "evidence_zh": val.get("evidence_zh") or ("无" if value == "未提及" else ""),
                "confidence": val.get("confidence", "medium"),
                "flags": val.get("flags", []) or [],
            }
        item["dimensions"] = dims
        item["context"] = ctx
        item["other"] = item.get("other") if isinstance(item.get("other"), dict) else {"T": "0", "R": "无"}
        item["need_review"] = bool(item.get("need_review"))
        item["review_flags"] = item.get("review_flags", []) or []
        normalized.append(item)
    missing = expected_ids - seen
    if missing:
        errors.append(f"missing review ids: {', '.join(sorted(missing)[:10])}")
    return {"reviews": normalized}, errors


def merge_final_labels(project_id: str, batch_id: str, result: dict, usage: dict):
    path = project_dir(project_id) / "final_labels.json"
    existing = read_json(path, [])
    existing = [r for r in existing if r.get("batch_id") != batch_id]
    for item in result.get("reviews", []):
        item["batch_id"] = batch_id
        item["usage"] = usage
        existing.append(item)
    write_json(path, existing)


def build_export(project_id: str):
    project = get_project(project_id)
    reviews = read_json(project_dir(project_id) / "reviews.json", [])
    batches = read_json(project_dir(project_id) / "batches.json", [])
    atomic = read_json(project_dir(project_id) / "atomic_results.json", [])
    candidates = read_json(project_dir(project_id) / "dimension_candidates.json", {})
    dimension_model = read_json(project_dir(project_id) / "dimension_model.json", {})
    dimension_stats = project_dimension_stats(project_id)
    analysis = read_json(project_dir(project_id) / "analysis_summary.json", {})
    locked = rules_for_project(project_id)
    final_labels = read_json(project_dir(project_id) / "final_labels.json", [])
    final_by_review = {row.get("review_id"): row for row in final_labels}
    decision_dims = [d.get("name_zh", "") for d in locked.get("decision_dimensions", []) if d.get("name_zh")]
    context_fields = [c.get("name_zh", "") for c in locked.get("context_fields", []) if c.get("name_zh")]

    wb = Workbook()
    ws = wb.active
    ws.title = "总览"
    rows = [
        ["项目", project.get("name", "") if project else ""],
        ["产品品类", project.get("category", "") if project else ""],
        ["评论数", len(reviews)],
        ["批次数", len(batches)],
        ["最小语义标签记录数", sum(len(x.get("atomic_tags", [])) for x in atomic)],
        ["NeedReview", sum(1 for x in atomic if x.get("need_review"))],
        ["导出时间", now_iso()],
    ]
    for r in rows:
        ws.append(r)

    ws = wb.create_sheet("review打标宽表")
    headers = ["Review序号", "ReviewID", "型号", "星级", "Review原文", "Review中文翻译", "Review链接", "NeedReview", "复核标记"]
    for name in context_fields:
        headers.extend([name, f"{name}证据"])
    for name in decision_dims:
        headers.extend([f"{name}T", f"{name}R", f"{name}原子属性"])
    headers.extend(["其他T", "其他R"])
    ws.append(headers)
    for r in reviews:
        label = final_by_review.get(r.get("review_id"), {})
        row = [
            r.get("seq"),
            r.get("review_id"),
            r.get("model"),
            r.get("star"),
            r.get("review_original"),
            r.get("review_translation_zh"),
            r.get("review_link"),
            label.get("need_review", ""),
            "；".join(label.get("review_flags", []) or []),
        ]
        for name in context_fields:
            item = (label.get("context", {}) or {}).get(name, {})
            row.extend([item.get("value_zh", "未提及"), item.get("evidence_zh", "无")])
        for name in decision_dims:
            item = (label.get("dimensions", {}) or {}).get(name, {})
            row.extend([item.get("T", "0"), item.get("R", "无"), item.get("atomic_value_zh", "无")])
        other = label.get("other", {}) or {}
        row.extend([other.get("T", "0"), other.get("R", "无")])
        ws.append(row)

    ws = wb.create_sheet("评论明细")
    ws.append(["Review序号", "ReviewID", "型号", "星级", "Review原文", "Review中文翻译", "Review链接"])
    for r in reviews:
        ws.append([r.get("seq"), r.get("review_id"), r.get("model"), r.get("star"), r.get("review_original"), r.get("review_translation_zh"), r.get("review_link")])

    ws = wb.create_sheet("最小语义标签明细")
    ws.append(["ReviewID", "批次", "最小语义标签", "初始倾向", "用途标记", "原文证据", "置信度", "NeedReview", "复核标记"])
    for row in atomic:
        for tag in row.get("atomic_tags", []) or []:
            ws.append(
                [
                    row.get("review_id"),
                    row.get("batch_id"),
                    tag.get("atomic_tag_zh"),
                    tag.get("sentiment"),
                    "、".join(tag.get("usage_marks", []) or []),
                    tag.get("evidence_original"),
                    tag.get("confidence"),
                    row.get("need_review"),
                    "；".join(row.get("review_flags", []) or []),
                ]
            )

    ws = wb.create_sheet("Drop记录")
    ws.append(["ReviewID", "批次", "Drop文本", "Drop原因"])
    for row in atomic:
        for drop in row.get("drop_records", []) or []:
            ws.append([row.get("review_id"), row.get("batch_id"), drop.get("text"), drop.get("drop_reason")])

    ws = wb.create_sheet("候选语义池")
    ws.append(["类型", "最小语义标签", "提及次数"])
    for item in candidates.get("product_atomic_pool", []):
        ws.append(["产品表现信号池", item["atomic_tag"], item["count"]])
    for item in candidates.get("context_atomic_pool", []):
        ws.append(["背景信号池", item["atomic_tag"], item["count"]])

    ws = wb.create_sheet("维度提及率")
    ws.append(["阶段", "类型", "维度/字段", "提及Review数", "提及率", "P数", "N数", "M数", "排序口径"])
    for stage, block in [("草案来源", dimension_stats.get("source", {})), ("最终打标", dimension_stats.get("final", {}))]:
        for item in block.get("decision", []) or []:
            ws.append(
                [
                    stage,
                    "购买决策维度",
                    item.get("name_zh"),
                    item.get("mention_count"),
                    item.get("mention_rate"),
                    item.get("p_count", ""),
                    item.get("n_count", ""),
                    item.get("m_count", ""),
                    item.get("sort_basis", ""),
                ]
            )
        for item in block.get("context", []) or []:
            ws.append(
                [
                    stage,
                    "Context",
                    item.get("name_zh"),
                    item.get("mention_count"),
                    item.get("mention_rate"),
                    "",
                    "",
                    "",
                    item.get("sort_basis", ""),
                ]
            )

    ws = wb.create_sheet("NeedReview复核清单")
    ws.append(["阶段", "ReviewID", "原因", "证据", "Review原文"])
    for item in need_review_items(project_id, limit=10000):
        ws.append([item.get("stage"), item.get("review_id"), item.get("reason"), item.get("evidence"), item.get("review_original")])

    ws = wb.create_sheet("决策细分与切入口")
    ws.append(["父维度", "细分维度/原子属性", "提及Review数", "提及率", "P数", "N数", "M数", "优化切入口", "证据示例", "ReviewID"])
    for item in analysis.get("decision_subdimensions", []) or []:
        ws.append(
            [
                item.get("parent_dimension"),
                item.get("subdimension_zh"),
                item.get("mention_count"),
                item.get("mention_rate"),
                item.get("p_count"),
                item.get("n_count"),
                item.get("m_count"),
                item.get("optimization_entry_zh"),
                "；".join(item.get("evidence_examples", []) or []),
                "；".join(item.get("review_ids", []) or []),
            ]
        )

    ws = wb.create_sheet("Context细分与洞察")
    ws.append(["Context字段", "细分值/原子属性", "提及Review数", "提及率", "分析用途", "证据示例", "ReviewID"])
    for item in analysis.get("context_subdimensions", []) or []:
        ws.append(
            [
                item.get("context_field"),
                item.get("subdimension_zh"),
                item.get("mention_count"),
                item.get("mention_rate"),
                item.get("analysis_use_zh"),
                "；".join(item.get("evidence_examples", []) or []),
                "；".join(item.get("review_ids", []) or []),
            ]
        )

    ws = wb.create_sheet("维度规则草案")
    ws.append(["类型", "名称", "定义", "P规则/证据要求", "N规则", "M规则", "0规则", "边界", "来源最小语义标签", "运营用途"])
    for item in dimension_model.get("decision_dimensions", []):
        ws.append(
            [
                "购买决策维度",
                item.get("name_zh"),
                item.get("definition_zh"),
                item.get("p_rule_zh"),
                item.get("n_rule_zh"),
                item.get("m_rule_zh"),
                item.get("zero_rule_zh"),
                item.get("boundary_zh"),
                "；".join(item.get("source_atomic_tags", []) or []),
                item.get("listing_use_zh"),
            ]
        )
    for item in dimension_model.get("context_fields", []):
        ws.append(
            [
                "Context字段",
                item.get("name_zh"),
                item.get("definition_zh"),
                item.get("evidence_required_zh"),
                "",
                "",
                "",
                item.get("boundary_zh"),
                "；".join(item.get("source_atomic_tags", []) or []),
                item.get("analysis_use_zh"),
            ]
        )

    ws = wb.create_sheet("批次状态")
    ws.append(["批次", "起始序号", "结束序号", "评论数", "状态", "模型", "错误"])
    for b in batches:
        ws.append([b.get("id"), b.get("start_seq"), b.get("end_seq"), b.get("review_count"), b.get("status"), b.get("model"), b.get("error")])

    for ws in wb.worksheets:
        ws.freeze_panes = "A2" if ws.max_row > 1 else None
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Microsoft YaHei", size=10, color=cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        if ws.max_column >= 5:
            ws.column_dimensions["E"].width = 70
        if ws.max_column >= 6:
            ws.column_dimensions["F"].width = 50

    out = EXPORT_DIR / f"{project_id}-voc-mvp-export.xlsx"
    wb.save(out)
    return out


class Handler(SimpleHTTPRequestHandler):
    server_version = "VOCWebTool/0.1"

    def translate_path(self, path):
        parsed = urlparse(path)
        if parsed.path == "/":
            return str(STATIC_DIR / "index.html")
        if parsed.path.startswith("/static/"):
            return str(STATIC_DIR / parsed.path.replace("/static/", "", 1))
        return str(STATIC_DIR / parsed.path.lstrip("/"))

    def log_message(self, fmt, *args):
        print("[%s] %s" % (now_iso(), fmt % args))

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if not has_access(self):
            if path.startswith("/api/"):
                send_json(self, {"error": "unauthorized"}, 401)
                return
            send_html(self, login_page())
            return
        if path == "/api/health":
            send_json(self, {"ok": True, "time": now_iso()})
            return
        if path == "/api/config":
            send_json(
                self,
                {
                    "has_deepseek_key": bool(os.environ.get("DEEPSEEK_API_KEY", "")),
                    "fast_model": DEFAULT_FAST_MODEL,
                    "accurate_model": DEFAULT_ACCURATE_MODEL,
                },
            )
            return
        if path == "/api/projects":
            projects = load_projects()
            enriched = [{**p, "stats": project_stats(p["id"])} for p in projects]
            send_json(self, {"projects": enriched})
            return
        m = re.match(r"^/api/projects/([^/]+)$", path)
        if m:
            project_id = m.group(1)
            project = get_project(project_id)
            if not project:
                send_json(self, {"error": "project_not_found"}, 404)
                return
            send_json(
                self,
                {
                    "project": {**project, "stats": project_stats(project_id)},
                    "reviews": read_json(project_dir(project_id) / "reviews.json", [])[:20],
                    "batches": read_json(project_dir(project_id) / "batches.json", []),
                    "atomic_results": read_json(project_dir(project_id) / "atomic_results.json", [])[:30],
                    "dimension_candidates": read_json(project_dir(project_id) / "dimension_candidates.json", {}),
                    "dimension_model": read_json(project_dir(project_id) / "dimension_model.json", {}),
                    "locked_dimensions": read_locked_rules(project_id),
                    "dimension_stats": project_dimension_stats(project_id),
                    "need_review_items": need_review_items(project_id),
                    "final_labels": read_json(project_dir(project_id) / "final_labels.json", [])[:30],
                    "analysis_summary": read_json(project_dir(project_id) / "analysis_summary.json", {}),
                },
            )
            return
        m = re.match(r"^/api/projects/([^/]+)/export$", path)
        if m:
            out = build_export(m.group(1))
            send_file(self, out, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            return
        return super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/api/access":
            data = body_json(self)
            if ACCESS_CODE and data.get("code") == ACCESS_CODE:
                payload = json.dumps({"ok": True}).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(payload)))
                self.send_header("Set-Cookie", "voc_access=1; Path=/; HttpOnly; SameSite=Lax; Max-Age=2592000")
                self.end_headers()
                self.wfile.write(payload)
            else:
                send_json(self, {"error": "invalid_access_code"}, 401)
            return
        if not has_access(self):
            send_json(self, {"error": "unauthorized"}, 401)
            return
        if path == "/api/projects":
            data = body_json(self)
            projects = load_projects()
            project_id = uuid.uuid4().hex[:10]
            project = {
                "id": project_id,
                "name": data.get("name", "未命名VOC项目"),
                "category": data.get("category", ""),
                "description": data.get("description", ""),
                "batch_size": int(data.get("batch_size") or 25),
                "fast_model": data.get("fast_model") or DEFAULT_FAST_MODEL,
                "accurate_model": data.get("accurate_model") or DEFAULT_ACCURATE_MODEL,
                "stage": "created",
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
            projects.insert(0, project)
            save_projects(projects)
            project_dir(project_id).mkdir(parents=True, exist_ok=True)
            send_json(self, {"project": project})
            return

        m = re.match(r"^/api/projects/([^/]+)/reviews$", path)
        if m:
            project_id = m.group(1)
            if not get_project(project_id):
                send_json(self, {"error": "project_not_found"}, 404)
                return
            form = FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST"})
            file_item = form["file"] if "file" in form else None
            if file_item is None or not getattr(file_item, "filename", ""):
                send_json(self, {"error": "missing_file"}, 400)
                return
            raw = file_item.file.read()
            name = file_item.filename.lower()
            if name.endswith(".xlsx"):
                reviews, parse_info = parse_reviews_from_xlsx(raw)
            elif name.endswith(".csv"):
                reviews, parse_info = parse_reviews_from_csv(raw)
            else:
                send_json(self, {"error": "unsupported_file_type"}, 400)
                return
            if not reviews:
                send_json(
                    self,
                    {
                        "error": "no_reviews_recognized",
                        "detail": "没有识别到 Review。请确认文件里有“Review原文 / 正文 / Review Body / Review Text / 评论内容”等正文列，或至少有一列包含完整评论文本。",
                        "parse_info": parse_info,
                    },
                    400,
                )
                return
            reset_after_reviews_upload(project_id)
            write_json(project_dir(project_id) / "reviews.json", reviews)
            project = update_project(project_id, {"stage": "reviews_imported"})
            send_json(self, {"project": project, "stats": project_stats(project_id), "sample": reviews[:10], "parse_info": parse_info})
            return

        m = re.match(r"^/api/projects/([^/]+)/batches$", path)
        if m:
            project_id = m.group(1)
            data = body_json(self)
            project = get_project(project_id)
            if not project:
                send_json(self, {"error": "project_not_found"}, 404)
                return
            size = int(data.get("batch_size") or project.get("batch_size") or 25)
            batches = make_batches(project_id, size)
            update_project(project_id, {"stage": "batched", "batch_size": size})
            send_json(self, {"batches": batches, "stats": project_stats(project_id)})
            return

        m = re.match(r"^/api/projects/([^/]+)/propose-dimensions$", path)
        if m:
            project_id = m.group(1)
            project = get_project(project_id)
            if not project:
                send_json(self, {"error": "project_not_found"}, 404)
                return
            data = body_json(self)
            mock = bool(data.get("mock"))
            model = data.get("model") or project.get("accurate_model") or DEFAULT_ACCURATE_MODEL
            api_key = self.headers.get("X-DeepSeek-Key") or os.environ.get("DEEPSEEK_API_KEY", "")
            if not mock and not api_key:
                send_json(self, {"error": "missing_deepseek_key"}, 400)
                return
            try:
                draft = generate_dimension_model(project_id, api_key, model, mock)
                send_json(self, {"status": "ok", "dimension_model": draft, "stats": project_stats(project_id)})
            except Exception as e:
                send_json(self, {"error": "dimension_model_failed", "detail": str(e)}, 500)
            return

        m = re.match(r"^/api/projects/([^/]+)/dimensions$", path)
        if m:
            project_id = m.group(1)
            if not get_project(project_id):
                send_json(self, {"error": "project_not_found"}, 404)
                return
            data = body_json(self)
            payload = data.get("dimensions") if isinstance(data.get("dimensions"), dict) else data
            decision = payload.get("decision_dimensions", [])
            context = payload.get("context_fields", [])
            if not isinstance(decision, list) or not isinstance(context, list):
                send_json(self, {"error": "invalid_dimensions_payload"}, 400)
                return
            reset_after_dimension_change(project_id)
            payload["locked_at"] = now_iso()
            write_json(project_dir(project_id) / "locked_dimensions.json", payload)
            update_project(project_id, {"stage": "dimensions_locked"})
            send_json(self, {"status": "ok", "locked_dimensions": payload, "stats": project_stats(project_id)})
            return

        m = re.match(r"^/api/projects/([^/]+)/unlock-dimensions$", path)
        if m:
            project_id = m.group(1)
            if not get_project(project_id):
                send_json(self, {"error": "project_not_found"}, 404)
                return
            archive_and_remove(project_id, ["locked_dimensions.json", "final_labels.json", "analysis_summary.json"])
            update_project(project_id, {"stage": "dimension_draft_ready"})
            send_json(self, {"status": "ok", "message": "已解除锁定。最终打标和分析结果已归档，因为它们依赖旧规则。", "stats": project_stats(project_id)})
            return

        m = re.match(r"^/api/projects/([^/]+)/analysis$", path)
        if m:
            project_id = m.group(1)
            if not get_project(project_id):
                send_json(self, {"error": "project_not_found"}, 404)
                return
            final_labels = read_json(project_dir(project_id) / "final_labels.json", [])
            if not final_labels:
                send_json(self, {"error": "missing_final_labels", "detail": "请先完成第 6 步最终打标，再生成细分维度和优化切入口。"}, 400)
                return
            summary = build_analysis_summary(project_id)
            send_json(self, {"status": "ok", "analysis_summary": summary, "stats": project_stats(project_id)})
            return

        m = re.match(r"^/api/projects/([^/]+)/batches/([^/]+)/extract-atomic$", path)
        if m:
            project_id, batch_id = m.group(1), m.group(2)
            project = get_project(project_id)
            if not project:
                send_json(self, {"error": "project_not_found"}, 404)
                return
            data = body_json(self)
            model = data.get("model") or project.get("fast_model") or DEFAULT_FAST_MODEL
            mock = bool(data.get("mock"))
            api_key = self.headers.get("X-DeepSeek-Key") or os.environ.get("DEEPSEEK_API_KEY", "")
            batch, reviews = batch_reviews(project_id, batch_id)
            if not batch:
                send_json(self, {"error": "batch_not_found"}, 404)
                return
            if not reviews:
                send_json(self, {"error": "empty_batch"}, 400)
                return
            if not mock and not api_key:
                send_json(self, {"error": "missing_deepseek_key"}, 400)
                return
            set_batch_status(project_id, batch_id, status="running", model=model, started_at=now_iso(), error="")
            started = time.time()
            try:
                if mock:
                    result, usage = mock_atomic_extract(reviews)
                else:
                    result, usage = deepseek_chat(api_key, model, atomic_prompt(project, reviews))
                errors = validate_atomic_result(result, {r["review_id"] for r in reviews})
                merge_atomic_results(project_id, batch_id, result, usage)
                propose_dimensions(project_id)
                set_batch_status(
                    project_id,
                    batch_id,
                    status="done" if not errors else "done_with_warnings",
                    finished_at=now_iso(),
                    error="；".join(errors[:5]),
                )
                send_json(
                    self,
                    {
                        "status": "ok",
                        "validation_errors": errors,
                        "usage": usage,
                        "duration_sec": round(time.time() - started, 1),
                        "stats": project_stats(project_id),
                    },
                )
            except Exception as e:
                set_batch_status(project_id, batch_id, status="failed", finished_at=now_iso(), error=str(e)[:1200])
                send_json(self, {"error": "extract_failed", "detail": str(e)}, 500)
            return

        m = re.match(r"^/api/projects/([^/]+)/batches/([^/]+)/label-final$", path)
        if m:
            project_id, batch_id = m.group(1), m.group(2)
            project = get_project(project_id)
            if not project:
                send_json(self, {"error": "project_not_found"}, 404)
                return
            rules = rules_for_project(project_id)
            if not rules.get("decision_dimensions") and not rules.get("context_fields"):
                send_json(self, {"error": "missing_locked_dimensions"}, 400)
                return
            data = body_json(self)
            model = data.get("model") or project.get("accurate_model") or DEFAULT_ACCURATE_MODEL
            mock = bool(data.get("mock"))
            api_key = self.headers.get("X-DeepSeek-Key") or os.environ.get("DEEPSEEK_API_KEY", "")
            batch, reviews = batch_reviews(project_id, batch_id)
            if not batch:
                send_json(self, {"error": "batch_not_found"}, 404)
                return
            if not reviews:
                send_json(self, {"error": "empty_batch"}, 400)
                return
            if not mock and not api_key:
                send_json(self, {"error": "missing_deepseek_key"}, 400)
                return
            set_batch_status(project_id, batch_id, status="final_running", model=model, started_at=now_iso(), error="")
            started = time.time()
            try:
                review_ids = {r["review_id"] for r in reviews}
                if mock:
                    result = {
                        "reviews": [
                            {
                                "review_id": r["review_id"],
                                "dimensions": {},
                                "context": {},
                                "other": {"T": "0", "R": "无"},
                                "need_review": True,
                                "review_flags": ["模拟最终打标，不作为正式结果"],
                            }
                            for r in reviews
                        ]
                    }
                    usage = {"mock": True}
                else:
                    atomic_rows = atomic_for_review(project_id, review_ids)
                    result, usage = deepseek_chat(api_key, model, final_label_prompt(project, rules, reviews, atomic_rows), max_tokens=24000, temperature=0.08)
                normalized, errors = normalize_final_result(result, review_ids, rules)
                merge_final_labels(project_id, batch_id, normalized, usage)
                set_batch_status(
                    project_id,
                    batch_id,
                    status="final_done" if not errors else "final_done_with_warnings",
                    finished_at=now_iso(),
                    error="；".join(errors[:5]),
                )
                update_project(project_id, {"stage": "final_labeling"})
                send_json(
                    self,
                    {
                        "status": "ok",
                        "validation_errors": errors,
                        "usage": usage,
                        "duration_sec": round(time.time() - started, 1),
                        "stats": project_stats(project_id),
                    },
                )
            except Exception as e:
                set_batch_status(project_id, batch_id, status="final_failed", finished_at=now_iso(), error=str(e)[:1200])
                send_json(self, {"error": "final_label_failed", "detail": str(e)}, 500)
            return

        send_json(self, {"error": "not_found"}, 404)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if not has_access(self):
            send_json(self, {"error": "unauthorized"}, 401)
            return
        m = re.match(r"^/api/projects/([^/]+)$", path)
        if m:
            project = delete_project(m.group(1))
            if not project:
                send_json(self, {"error": "project_not_found"}, 404)
                return
            send_json(self, {"status": "ok", "deleted_project": project})
            return
        send_json(self, {"error": "not_found"}, 404)


def main():
    ensure_dirs()
    host = os.environ.get("VOC_HOST", "127.0.0.1")
    port = int(os.environ.get("VOC_PORT") or os.environ.get("PORT") or "8787")
    httpd = ThreadingHTTPServer((host, port), Handler)
    print(f"VOC Web Tool running at http://{host}:{port}")
    print("Set DEEPSEEK_API_KEY in environment before running formal labeling.")
    httpd.serve_forever()


if __name__ == "__main__":
    main()
